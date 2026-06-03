#!/usr/bin/env python3
"""
Colorado Small — IP clicks matched to targeted employers, rendered as an
interactive, filterable GIS dashboard.

Inputs  (two tabs of Colorado Small.xlsx):
  - "Colorado Employers"     : Company name + mailing address (the 520 targeted employers)
  - "Colorado Small Clicks"  : IP address + click count (the delivered clicks)

Pipeline:
  1. Read both tabs.
  2. Geocode employer addresses  -> lat/lon  (US Census batch geocoder, free, no key).
  3. Geolocate clicking IPs       -> lat/lon + org  (ipinfo.io batch API).
  4. Match each IP to the NEAREST employer (haversine); record distance + confidence.
  5. Emit a self-contained Leaflet HTML dashboard (map + filterable table).
"""

import csv
import io
import json
import math
import urllib.request
import urllib.parse
import ssl
from collections import defaultdict

# macOS python.org builds often lack a CA bundle; prefer certifi, fall back to unverified.
try:
    import certifi
    SSL_CTX = ssl.create_default_context(cafile=certifi.where())
except Exception:
    SSL_CTX = ssl.create_default_context()
    SSL_CTX.check_hostname = False
    SSL_CTX.verify_mode = ssl.CERT_NONE

import openpyxl
from dashboard_template import HTML_TEMPLATE

XLSX_PATH   = '/Users/scootervineburgh/Desktop/Kathi/Colorado Small.xlsx'
OUTPUT_PATH = '/Users/scootervineburgh/Desktop/Kathi/colorado_click_map.html'

IPINFO_TOKEN     = '359b803b357850'
IPINFO_BATCH_URL = f'https://ipinfo.io/batch?token={IPINFO_TOKEN}'
IPINFO_BATCH_SIZE = 1000

CENSUS_BATCH_URL = 'https://geocoding.geo.census.gov/geocoder/locations/addressbatch'
CENSUS_ONELINE   = 'https://geocoding.geo.census.gov/geocoder/locations/onelineaddress'

# ── 1. Read both tabs ─────────────────────────────────────────────────────────
print("Reading workbook...")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

emp_ws = wb['Colorado Employers']
employers = []
for i, row in enumerate(emp_ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    name, addr, city, state, zip5, zip4 = (list(row) + [None] * 6)[:6]
    if not name:
        continue
    zip_str = ''
    if zip5 not in (None, ''):
        zip_str = str(zip5).split('.')[0].zfill(5)
    zip4_str = ''
    if zip4 not in (None, ''):
        zip4_str = str(zip4).split('.')[0].zfill(4)
    employers.append({
        'id': len(employers),
        'name': str(name).strip(),              # Contact Company name
        'address': str(addr).strip() if addr else '',   # Contact Address
        'city': str(city).strip() if city else '',      # Contact City
        'state': 'CO',                          # normalized for geocoding
        'state_orig': str(state).strip() if state else '',  # Contact State (verbatim)
        'zip': zip_str,                         # Contact Zip
        'zip4': zip4_str,                       # Contact Zip4
    })
print(f"  {len(employers)} employers")

clicks_ws = wb['Colorado Small Clicks']
ip_clicks = defaultdict(int)
for i, row in enumerate(clicks_ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    ip = row[0]
    clk = row[1]
    if ip is None:
        continue
    ip = str(ip).strip()
    try:
        clk = int(clk) if clk is not None else 0
    except (ValueError, TypeError):
        clk = 0
    if ip:
        ip_clicks[ip] += clk
unique_ips = list(ip_clicks.keys())
print(f"  {len(unique_ips)} unique IPs, {sum(ip_clicks.values())} total clicks")

# ── 2. Geocode employer addresses via Census batch geocoder ───────────────────
print("Geocoding employer addresses (US Census batch)...")

def census_batch(records):
    """records: list of (id, street, city, state, zip). Returns {id: (lat, lon)}."""
    buf = io.StringIO()
    w = csv.writer(buf)
    for rid, street, city, state, zp in records:
        w.writerow([rid, street, city, state, zp])
    payload = buf.getvalue().encode()

    boundary = '----coloradoBatchBoundary'
    body = io.BytesIO()
    def part(name, value, filename=None, ctype=None):
        body.write(f'--{boundary}\r\n'.encode())
        if filename:
            body.write(f'Content-Disposition: form-data; name="{name}"; filename="{filename}"\r\n'.encode())
            body.write(f'Content-Type: {ctype}\r\n\r\n'.encode())
            body.write(value)
            body.write(b'\r\n')
        else:
            body.write(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
            body.write(f'{value}\r\n'.encode())
    part('benchmark', 'Public_AR_Current')
    part('addressFile', payload, filename='addr.csv', ctype='text/csv')
    body.write(f'--{boundary}--\r\n'.encode())

    req = urllib.request.Request(
        CENSUS_BATCH_URL, data=body.getvalue(),
        headers={'Content-Type': f'multipart/form-data; boundary={boundary}'})
    out = {}
    with urllib.request.urlopen(req, timeout=120, context=SSL_CTX) as resp:
        text = resp.read().decode('utf-8', errors='replace')
    for line in csv.reader(io.StringIO(text)):
        # id, input, match_status, match_type, matched_addr, "lon,lat", tigerid, side
        if len(line) < 6:
            continue
        rid = line[0]
        status = line[2]
        if status == 'Match' and line[5]:
            try:
                lon, lat = line[5].split(',')
                out[rid] = (float(lat), float(lon))
            except ValueError:
                pass
    return out

emp_coords = {}
BATCH = 1000
for start in range(0, len(employers), BATCH):
    chunk = employers[start:start + BATCH]
    recs = [(str(e['id']), e['address'], e['city'], e['state'], e['zip']) for e in chunk]
    try:
        res = census_batch(recs)
        emp_coords.update(res)
    except Exception as exc:
        print(f"  batch {start} error: {exc}")
    print(f"  geocoded {len(emp_coords)}/{len(employers)} employers")

# Fallback: one-line geocoder for any misses (ZIP-centroid quality is fine here)
missing = [e for e in employers if str(e['id']) not in emp_coords]
print(f"  {len(missing)} employers need fallback geocoding...")
for e in missing:
    oneline = f"{e['address']}, {e['city']}, {e['state']} {e['zip']}"
    url = f"{CENSUS_ONELINE}?address={urllib.parse.quote(oneline)}&benchmark=Public_AR_Current&format=json"
    try:
        with urllib.request.urlopen(url, timeout=20, context=SSL_CTX) as resp:
            data = json.loads(resp.read())
        matches = data['result']['addressMatches']
        if matches:
            c = matches[0]['coordinates']
            emp_coords[str(e['id'])] = (c['y'], c['x'])
    except Exception:
        pass
print(f"  total geocoded employers: {len(emp_coords)}/{len(employers)}")

# Stage 3: Nominatim (rural-friendly) then ZIP-centroid for stubborn addresses.
import re, time
def zip5(z):
    m = re.search(r'\d{5}', z or '')
    return m.group(0) if m else ''

emp_precision = {rid: 'rooftop' for rid in emp_coords}   # default; zip-centroid marked below
still_missing = [e for e in employers if str(e['id']) not in emp_coords]
print(f"  {len(still_missing)} still missing; trying Nominatim + ZIP centroid...")
for e in still_missing:
    rid, z = str(e['id']), zip5(e['zip'])
    # 3a. Nominatim structured search (rooftop/street precision)
    try:
        params = urllib.parse.urlencode({
            'street': e['address'], 'city': e['city'], 'state': 'CO',
            'postalcode': z, 'country': 'USA', 'format': 'json', 'limit': 1})
        req = urllib.request.Request(
            'https://nominatim.openstreetmap.org/search?' + params,
            headers={'User-Agent': 'IPMap-geocoder/1.0 (cyberpracticesolutions@gmail.com)'})
        with urllib.request.urlopen(req, timeout=25, context=SSL_CTX) as resp:
            arr = json.loads(resp.read())
        time.sleep(1.1)  # Nominatim usage policy: <=1 req/sec
        if arr:
            emp_coords[rid] = (float(arr[0]['lat']), float(arr[0]['lon']))
            emp_precision[rid] = 'rooftop'
            continue
    except Exception:
        pass
    # 3b. ZIP centroid (approximate — uses the address's own ZIP, so KS etc. resolve correctly)
    if z:
        try:
            with urllib.request.urlopen(f'https://api.zippopotam.us/us/{z}', timeout=15, context=SSL_CTX) as resp:
                d = json.loads(resp.read())
            pl = d['places'][0]
            emp_coords[rid] = (float(pl['latitude']), float(pl['longitude']))
            emp_precision[rid] = 'zip'
        except Exception:
            pass
print(f"  total geocoded employers now: {len(emp_coords)}/{len(employers)}")

for e in employers:
    coord = emp_coords.get(str(e['id']))
    e['lat'], e['lon'] = (coord if coord else (None, None))
    e['precision'] = emp_precision.get(str(e['id']), 'rooftop')

# ── 3. Geolocate IPs via ipinfo.io batch ──────────────────────────────────────
print("Geolocating IPs (ipinfo batch)...")
geo = {}
for start in range(0, len(unique_ips), IPINFO_BATCH_SIZE):
    batch = unique_ips[start:start + IPINFO_BATCH_SIZE]
    req = urllib.request.Request(
        IPINFO_BATCH_URL, data=json.dumps(batch).encode(),
        headers={'Content-Type': 'application/json'})
    try:
        with urllib.request.urlopen(req, timeout=60, context=SSL_CTX) as resp:
            results = json.loads(resp.read())
        for ip, entry in results.items():
            if isinstance(entry, dict) and 'loc' in entry:
                lat, lon = entry['loc'].split(',')
                org = entry.get('org', '')
                # strip leading "AS#### " from org string
                if org.startswith('AS'):
                    parts = org.split(' ', 1)
                    org = parts[1] if len(parts) > 1 else org
                geo[ip] = {
                    'lat': float(lat), 'lon': float(lon),
                    'city': entry.get('city', ''), 'region': entry.get('region', ''),
                    'org': org,
                }
    except Exception as exc:
        print(f"  batch {start} error: {exc}")
    print(f"  geolocated {len(geo)}/{len(unique_ips)} IPs")

# ── 4. Match each IP to nearest employer ──────────────────────────────────────
print("Matching IPs to nearest employer...")
def haversine(a, b, c, d):
    R = 6371.0
    p1, p2 = math.radians(a), math.radians(c)
    dphi = math.radians(c - a)
    dl = math.radians(d - b)
    h = math.sin(dphi/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2*R*math.asin(math.sqrt(h))

geo_emps = [e for e in employers if e['lat'] is not None]

def confidence(km):
    if km is None:        return 'none'
    if km <= 2:           return 'high'
    if km <= 10:          return 'medium'
    if km <= 40:          return 'low'
    return 'very-low'

points = []
for ip in unique_ips:
    g = geo.get(ip)
    if not g:
        continue
    best, best_km = None, None
    for e in geo_emps:
        km = haversine(g['lat'], g['lon'], e['lat'], e['lon'])
        if best_km is None or km < best_km:
            best, best_km = e, km
    conf = confidence(best_km)
    emp_prec = best['precision'] if best else 'rooftop'
    # ZIP-centroid employers are only neighborhood-accurate; don't claim a High match to them.
    if emp_prec == 'zip' and conf == 'high':
        conf = 'medium'
    points.append({
        'ip': ip,
        'clicks': ip_clicks[ip],
        'lat': g['lat'], 'lon': g['lon'],
        'city': g['city'], 'region': g['region'], 'org': g['org'],
        'employer': best['name'] if best else '',
        'emp_id': best['id'] if best else None,
        'emp_addr': f"{best['address']}, {best['city']}, CO {best['zip']}" if best else '',
        'emp_address': best['address'] if best else '',
        'emp_city': best['city'] if best else '',
        'emp_state': (best['state_orig'] or 'Colorado') if best else '',
        'emp_zip': best['zip'] if best else '',
        'emp_zip4': best['zip4'] if best else '',
        'emp_lat': best['lat'] if best else None,
        'emp_lon': best['lon'] if best else None,
        'emp_precision': emp_prec,
        'dist_km': round(best_km, 2) if best_km is not None else None,
        'dist_mi': round(best_km * 0.621371, 2) if best_km is not None else None,
        'conf': conf,
    })

points.sort(key=lambda p: -p['clicks'])
matched = sum(1 for p in points if p['conf'] in ('high', 'medium'))
print(f"  {len(points)} mapped IPs, {matched} high/medium-confidence employer matches")

# Per-employer rollup — seed ALL employers (so the entire employers tab is captured),
# then add high/medium-confidence matched clicks.
emp_agg = {}
for e in employers:
    emp_agg[e['id']] = {
        'employer': e['name'], 'clicks': 0, 'ips': 0,
        'address': e['address'], 'city': e['city'],
        'state': e['state_orig'] or 'Colorado', 'zip': e['zip'], 'zip4': e['zip4'],
    }
for p in points:
    if p['emp_id'] is not None and p['conf'] in ('high', 'medium'):
        a = emp_agg[p['emp_id']]
        a['clicks'] += p['clicks']
        a['ips'] += 1
emp_summary = sorted(emp_agg.values(), key=lambda x: (-x['clicks'], x['employer']))

# ── 5. Emit HTML ──────────────────────────────────────────────────────────────
print("Writing HTML dashboard...")
total_clicks = sum(ip_clicks.values())
stats = {
    'total_clicks': total_clicks,
    'unique_ips': len(unique_ips),
    'mapped_ips': len(points),
    'employers': len(employers),
    'geocoded_employers': len(geo_emps),
    'matched': matched,
}
html = (HTML_TEMPLATE
        .replace('/*__POINTS__*/', json.dumps(points))
        .replace('/*__EMPLOYERS__*/', json.dumps([e for e in employers if e['lat'] is not None]))
        .replace('/*__EMP_SUMMARY__*/', json.dumps(emp_summary))
        .replace('/*__STATS__*/', json.dumps(stats)))
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"\nDone -> {OUTPUT_PATH}")
print(f"  {len(points)} click points, {len(geo_emps)} employers geocoded, {matched} matches")
