#!/usr/bin/env python3
"""
Apollo Small — IP clicks matched to targeted companies, rendered as an
interactive, filterable GIS dashboard.

Inputs (two tabs of Apollo Small.xlsx):
  - "Apollo Small Companies" : Company Name, Street Address, City, State, Country,
                               Zip Code, Zip Code Plus4, Website  (973 targeted companies)
  - "Apollo Small Clicks"    : IP, Clicks                          (the delivered clicks)

Matching (best signal wins):
  1. DOMAIN   — IP reverse-DNS hostname domain == company website domain
                (verified, location-independent; highest confidence).
  2. ORG/ISP  — IP network owner (ipinfo org) matches a company name.
  3. GEO      — nearest geocoded company by haversine distance (confidence by distance).

Geocoding: US Census batch -> Nominatim -> ZIP centroid (national scope).
IP geolocation: ipinfo.io batch + local reverse DNS.
"""

import csv
import io
import json
import math
import re
import socket
import time
import urllib.request
import urllib.parse
import ssl
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor

try:
    import certifi
    SSL_CTX = ssl.create_default_context(cafile=certifi.where())
except Exception:
    SSL_CTX = ssl.create_default_context()
    SSL_CTX.check_hostname = False
    SSL_CTX.verify_mode = ssl.CERT_NONE

import openpyxl
from dashboard_template_apollo import HTML_TEMPLATE

XLSX_PATH   = '/Users/scootervineburgh/Desktop/Kathi/Apollo Small.xlsx'
OUTPUT_PATH = '/Users/scootervineburgh/Desktop/Kathi/apollo_click_map.html'

IPINFO_TOKEN     = '359b803b357850'
IPINFO_BATCH_URL = f'https://ipinfo.io/batch?token={IPINFO_TOKEN}'
CENSUS_BATCH_URL = 'https://geocoding.geo.census.gov/geocoder/locations/addressbatch'

STATE_ABBR = {
    'alabama':'AL','alaska':'AK','arizona':'AZ','arkansas':'AR','california':'CA',
    'colorado':'CO','connecticut':'CT','delaware':'DE','district of columbia':'DC',
    'florida':'FL','georgia':'GA','hawaii':'HI','idaho':'ID','illinois':'IL',
    'indiana':'IN','iowa':'IA','kansas':'KS','kentucky':'KY','louisiana':'LA',
    'maine':'ME','maryland':'MD','massachusetts':'MA','michigan':'MI','minnesota':'MN',
    'mississippi':'MS','missouri':'MO','montana':'MT','nebraska':'NE','nevada':'NV',
    'new hampshire':'NH','new jersey':'NJ','new mexico':'NM','new york':'NY',
    'north carolina':'NC','north dakota':'ND','ohio':'OH','oklahoma':'OK','oregon':'OR',
    'pennsylvania':'PA','rhode island':'RI','south carolina':'SC','south dakota':'SD',
    'tennessee':'TN','texas':'TX','utah':'UT','vermont':'VT','virginia':'VA',
    'washington':'WA','west virginia':'WV','wisconsin':'WI','wyoming':'WY',
    'puerto rico':'PR',
}

# ── 1. Read both tabs ─────────────────────────────────────────────────────────
print("Reading workbook...")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

def z(v, width):
    if v in (None, ''):
        return ''
    s = str(v).split('.')[0]
    return s.zfill(width) if s.isdigit() else s

comp_ws = wb['Apollo Small Companies']
companies = []
for i, row in enumerate(comp_ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    name, addr, city, state, country, zip5, zip4, website = (list(row) + [None] * 8)[:8]
    if not name:
        continue
    companies.append({
        'id': len(companies),
        'name': str(name).strip(),
        'address': str(addr).strip() if addr else '',
        'city': str(city).strip() if city else '',
        'state': str(state).strip() if state else '',
        'country': str(country).strip() if country else '',
        'zip': z(zip5, 5),
        'zip4': z(zip4, 4),
        'website': str(website).strip() if website else '',
    })
print(f"  {len(companies)} companies")

clicks_ws = wb['Apollo Small Clicks']
ip_clicks = defaultdict(int)
for i, row in enumerate(clicks_ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    ip = row[0]
    if ip is None:
        continue
    ip = str(ip).strip()
    try:
        clk = int(row[1]) if row[1] is not None else 0
    except (ValueError, TypeError):
        clk = 0
    if ip:
        ip_clicks[ip] += clk
unique_ips = list(ip_clicks.keys())
print(f"  {len(unique_ips)} unique IPs, {sum(ip_clicks.values())} total clicks")

# ── 2. Website -> registered domain index ─────────────────────────────────────
TWO_LEVEL = {'co.uk','com.au','co.nz','co.jp','com.br','co.in','co.za','org.uk',
             'net.au','org.au','ac.uk','gov.uk','com.mx','com.sg','com.hk'}
def reg_domain(host):
    host = (host or '').strip().lower().rstrip('.')
    if not host or '.' not in host:
        return ''
    parts = host.split('.')
    if len(parts) >= 3 and '.'.join(parts[-2:]) in TWO_LEVEL:
        return '.'.join(parts[-3:])
    return '.'.join(parts[-2:])

def site_domain(url):
    u = (url or '').strip().lower()
    u = re.sub(r'^[a-z]+://', '', u)
    u = u.split('/')[0].split('?')[0].split(':')[0]
    if u.startswith('www.'):
        u = u[4:]
    return reg_domain(u)

domain_index = {}
for c in companies:
    d = site_domain(c['website'])
    c['domain'] = d
    if d and d not in domain_index:
        domain_index[d] = c
print(f"  {len(domain_index)} unique company website domains")

# Normalized company names for org/ISP matching
SUFFIX = re.compile(r'\b(inc|incorporated|llc|l\.l\.c|corp|corporation|co|company|ltd|limited|'
                    r'plc|lp|llp|group|holdings|systems|technologies|technology|solutions|'
                    r'services|enterprises|industries|the|and)\b', re.I)
def norm_name(s):
    s = re.sub(r'[^a-z0-9 ]', ' ', (s or '').lower())
    s = SUFFIX.sub(' ', s)
    return re.sub(r'\s+', ' ', s).strip()

ISP_WORDS = ('comcast','verizon','at&t','att ','charter','spectrum','cox','centurylink',
             'lumen','t-mobile','sprint','frontier','windstream','google','amazon','microsoft',
             'cloudflare','akamai','digitalocean','linode','ovh','hetzner','telecom','communications',
             'broadband','cable','wireless','internet','hosting','datacenter','data center')
comp_by_norm = {}
for c in companies:
    nn = norm_name(c['name'])
    if len(nn) >= 5 and nn not in comp_by_norm:
        comp_by_norm[c['id']] = nn

# ── 3. Geocode companies (Census batch -> Nominatim -> ZIP centroid) ──────────
print("Geocoding companies...")

def census_batch(records):
    buf = io.StringIO()
    csv.writer(buf).writerows(records)
    payload = buf.getvalue().encode()
    boundary = '----apolloBatch'
    body = io.BytesIO()
    def part(name, value, filename=None, ctype=None):
        body.write(f'--{boundary}\r\n'.encode())
        if filename:
            body.write(f'Content-Disposition: form-data; name="{name}"; filename="{filename}"\r\n'.encode())
            body.write(f'Content-Type: {ctype}\r\n\r\n'.encode()); body.write(value); body.write(b'\r\n')
        else:
            body.write(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
            body.write(f'{value}\r\n'.encode())
    part('benchmark', 'Public_AR_Current')
    part('addressFile', payload, filename='addr.csv', ctype='text/csv')
    body.write(f'--{boundary}--\r\n'.encode())
    req = urllib.request.Request(CENSUS_BATCH_URL, data=body.getvalue(),
        headers={'Content-Type': f'multipart/form-data; boundary={boundary}'})
    out = {}
    with urllib.request.urlopen(req, timeout=120, context=SSL_CTX) as resp:
        text = resp.read().decode('utf-8', errors='replace')
    for line in csv.reader(io.StringIO(text)):
        if len(line) < 6:
            continue
        if line[2] == 'Match' and line[5]:
            try:
                lon, lat = line[5].split(',')
                out[line[0]] = (float(lat), float(lon))
            except ValueError:
                pass
    return out

coords, precision = {}, {}
batchable = [c for c in companies if c['address']]   # Census needs a street address
recs = [(str(c['id']), c['address'], c['city'],
         STATE_ABBR.get(c['state'].lower(), c['state']), c['zip']) for c in batchable]
for s in range(0, len(recs), 1000):
    try:
        res = census_batch(recs[s:s+1000])
        coords.update(res)
        for rid in res:
            precision[rid] = 'rooftop'
    except Exception as exc:
        print(f"  census batch {s} error: {exc}")
print(f"  Census matched {len(coords)}/{len(companies)}")

missing = [c for c in companies if str(c['id']) not in coords]
print(f"  {len(missing)} need Nominatim / ZIP centroid...")
for c in missing:
    rid = str(c['id'])
    abbr = STATE_ABBR.get(c['state'].lower(), c['state'])
    z5 = (re.search(r'\d{5}', c['zip']) or [None])
    z5 = z5.group(0) if hasattr(z5, 'group') else ''
    # Nominatim
    try:
        params = urllib.parse.urlencode({k: v for k, v in {
            'street': c['address'], 'city': c['city'], 'state': abbr,
            'postalcode': z5, 'country': c['country'] or 'USA',
            'format': 'json', 'limit': 1}.items() if v})
        req = urllib.request.Request('https://nominatim.openstreetmap.org/search?' + params,
            headers={'User-Agent': 'IPMap-geocoder/1.0 (cyberpracticesolutions@gmail.com)'})
        with urllib.request.urlopen(req, timeout=25, context=SSL_CTX) as resp:
            arr = json.loads(resp.read())
        time.sleep(1.1)
        if arr:
            coords[rid] = (float(arr[0]['lat']), float(arr[0]['lon']))
            precision[rid] = 'rooftop'
            continue
    except Exception:
        pass
    # ZIP centroid
    if z5:
        try:
            with urllib.request.urlopen(f'https://api.zippopotam.us/us/{z5}', timeout=15, context=SSL_CTX) as resp:
                d = json.loads(resp.read())
            pl = d['places'][0]
            coords[rid] = (float(pl['latitude']), float(pl['longitude']))
            precision[rid] = 'zip'
        except Exception:
            pass
print(f"  total geocoded companies: {len(coords)}/{len(companies)}")

for c in companies:
    xy = coords.get(str(c['id']))
    c['lat'], c['lon'] = (xy if xy else (None, None))
    c['precision'] = precision.get(str(c['id']), 'rooftop')

geo_comps = [c for c in companies if c['lat'] is not None]

# ── 4. Geolocate IPs (ipinfo) + reverse DNS ───────────────────────────────────
print("Geolocating IPs (ipinfo)...")
geo = {}
req = urllib.request.Request(IPINFO_BATCH_URL, data=json.dumps(unique_ips).encode(),
    headers={'Content-Type': 'application/json'})
try:
    with urllib.request.urlopen(req, timeout=60, context=SSL_CTX) as resp:
        results = json.loads(resp.read())
    for ip, e in results.items():
        if isinstance(e, dict) and 'loc' in e:
            lat, lon = e['loc'].split(',')
            org = e.get('org', '')
            if org.startswith('AS'):
                org = org.split(' ', 1)[1] if ' ' in org else org
            geo[ip] = {'lat': float(lat), 'lon': float(lon),
                       'city': e.get('city', ''), 'region': e.get('region', ''),
                       'org': org, 'hostname': e.get('hostname', '')}
except Exception as exc:
    print(f"  ipinfo error: {exc}")
print(f"  geolocated {len(geo)}/{len(unique_ips)}")

print("Reverse-DNS lookups...")
socket.setdefaulttimeout(3)
def rdns(ip):
    try:
        return ip, socket.gethostbyaddr(ip)[0].lower()
    except Exception:
        return ip, ''
with ThreadPoolExecutor(max_workers=20) as ex:
    host_map = dict(ex.map(rdns, unique_ips))
# prefer ipinfo hostname when local PTR is empty
for ip in unique_ips:
    if not host_map.get(ip) and geo.get(ip, {}).get('hostname'):
        host_map[ip] = geo[ip]['hostname'].lower()
n_host = sum(1 for v in host_map.values() if v)
print(f"  {n_host}/{len(unique_ips)} IPs have a hostname")

# ── 5. Match each IP ──────────────────────────────────────────────────────────
print("Matching IPs to companies...")
def haversine(a, b, c, d):
    R = 6371.0
    p1, p2 = math.radians(a), math.radians(c)
    h = math.sin(math.radians(c-a)/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(math.radians(d-b)/2)**2
    return 2*R*math.asin(math.sqrt(h))

def dist_conf(km):
    if km is None:  return 'none'
    if km <= 2:     return 'high'
    if km <= 10:    return 'medium'
    if km <= 40:    return 'low'
    return 'very-low'

def org_match(org):
    """Return a company whose name clearly appears in the IP's network-owner string."""
    no = norm_name(org)
    low = (org or '').lower()
    if not no or any(w in low for w in ISP_WORDS):
        return None
    for cid, nn in comp_by_norm.items():
        if len(nn) >= 6 and (nn in no or no in nn):
            return companies[cid]
    return None

points = []
for ip in unique_ips:
    g = geo.get(ip)
    if not g:
        continue
    host = host_map.get(ip, '')
    hdom = reg_domain(host)
    match, mtype, conf = None, 'geo', 'none'

    # 1. domain
    if hdom and hdom in domain_index:
        match, mtype, conf = domain_index[hdom], 'domain', 'high'
    # 2. org / ISP owner
    if match is None:
        om = org_match(g['org'])
        if om:
            match, mtype, conf = om, 'org', 'medium'
    # 3. geo nearest
    if match is None and geo_comps:
        best, best_km = None, None
        for c in geo_comps:
            km = haversine(g['lat'], g['lon'], c['lat'], c['lon'])
            if best_km is None or km < best_km:
                best, best_km = c, km
        match, mtype = best, 'geo'
        conf = dist_conf(best_km)
        if best and best['precision'] == 'zip' and conf == 'high':
            conf = 'medium'

    dist_km = None
    if match and match['lat'] is not None:
        dist_km = haversine(g['lat'], g['lon'], match['lat'], match['lon'])

    points.append({
        'ip': ip, 'clicks': ip_clicks[ip],
        'lat': g['lat'], 'lon': g['lon'],
        'city': g['city'], 'region': g['region'], 'org': g['org'], 'hostname': host,
        'company': match['name'] if match else '',
        'comp_id': match['id'] if match else None,
        'website': match['website'] if match else '',
        'comp_addr': ', '.join(filter(None, [match['address'], match['city'], match['state'],
                     '-'.join(filter(None, [match['zip'], match['zip4']]))])) if match else '',
        'comp_address': match['address'] if match else '',
        'comp_city': match['city'] if match else '',
        'comp_state': match['state'] if match else '',
        'comp_country': match['country'] if match else '',
        'comp_zip': match['zip'] if match else '',
        'comp_zip4': match['zip4'] if match else '',
        'comp_lat': match['lat'] if match else None,
        'comp_lon': match['lon'] if match else None,
        'comp_precision': match['precision'] if match else '',
        'match_type': mtype,
        'dist_km': round(dist_km, 2) if dist_km is not None else None,
        'dist_mi': round(dist_km * 0.621371, 2) if dist_km is not None else None,
        'conf': conf,
    })

points.sort(key=lambda p: -p['clicks'])
n_domain = sum(1 for p in points if p['match_type'] == 'domain')
n_org = sum(1 for p in points if p['match_type'] == 'org')
matched = sum(1 for p in points if p['match_type'] in ('domain', 'org') or p['conf'] in ('high', 'medium'))
print(f"  {len(points)} mapped IPs | {n_domain} domain, {n_org} org, "
      f"{matched} total high-quality matches")

# Per-company rollup — seed ALL companies, add domain/org/geo-high/medium clicks
comp_agg = {}
for c in companies:
    comp_agg[c['id']] = {
        'company': c['name'], 'website': c['website'], 'clicks': 0, 'ips': 0,
        'address': c['address'], 'city': c['city'], 'state': c['state'],
        'country': c['country'], 'zip': c['zip'], 'zip4': c['zip4'],
    }
for p in points:
    if p['comp_id'] is not None and (p['match_type'] in ('domain', 'org') or p['conf'] in ('high', 'medium')):
        a = comp_agg[p['comp_id']]
        a['clicks'] += p['clicks']
        a['ips'] += 1
comp_summary = sorted(comp_agg.values(), key=lambda x: (-x['clicks'], x['company']))

# ── 6. Emit HTML ──────────────────────────────────────────────────────────────
print("Writing dashboard...")
stats = {
    'total_clicks': sum(ip_clicks.values()),
    'unique_ips': len(unique_ips),
    'mapped_ips': len(points),
    'companies': len(companies),
    'geocoded_companies': len(geo_comps),
    'matched': matched,
    'domain': n_domain,
}
html = (HTML_TEMPLATE
        .replace('/*__POINTS__*/', json.dumps(points))
        .replace('/*__COMPANIES__*/', json.dumps(geo_comps))
        .replace('/*__COMP_SUMMARY__*/', json.dumps(comp_summary))
        .replace('/*__STATS__*/', json.dumps(stats)))
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"\nDone -> {OUTPUT_PATH}")
print(f"  {len(points)} click points, {len(geo_comps)} companies geocoded, "
      f"{n_domain} domain-verified, {matched} matched")
